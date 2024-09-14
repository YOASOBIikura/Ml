import numpy as np
import math
import openpyxl

# file_admet = openpyxl.load_workbook('ADMET.xlsx')
file_admet = openpyxl.load_workbook('ADMET.xlsx')
# file_admet_training_sheet = file_admet('training')
file_admet_training_sheet = file_admet.sheetnames


# store admet file data
file_admet_data, file_admet_index = [], []

for line_data in file_admet_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=6):
    temp = []
    for c in line_data:
        temp.append(c.value)
    file_admet_data.append(temp)

# admet rules, get fit data index, as row num
for step, item in enumerate(file_admet_data):
    if item.count(1) >= 3:
        file_admet_index.append(step)

file_molecular = openpyxl.load_workbook('Molecular_Descriptor.xlsx')
file_molecular_training_sheet = file_molecular('training')

# store molecular file training data and columns names
file_molecular_training_data, file_molecular_columns = [], []

# training data
for line_data in file_molecular_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=730):
    temp = []
    for c in line_data:
        temp.append(c.value)
    file_molecular_training_data.append(temp)

# columns names
for line_data in file_molecular_training_sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=730):
    temp = []
    for c in line_data:
        temp.append(c.value)
    file_molecular_columns.append(temp)

# variables 20
var_names_20 = []
with open('vars_20.txt', 'r', encoding='utf-8') as f:
    nn = f.readlines()
    for item in nn:
        var_names_20.append(item.split(':')[0])

# 20 vars index in columns
var_20_names_index = [file_molecular_columns[0].index(item) for item in var_names_20]

file_er = openpyxl.load_workbook('ER_activity.xlsx')
file_er_training_sheet = file_er('training')

# store er file training data
file_er_training_data = []

for line_data in file_er_training_sheet.iter_rows(min_row=2, max_row=1975, min_col=2, max_col=2):
    file_er_training_data.append([c.value for c in line_data])

# get data that fit rule admet
fit_admet_data = [file_molecular_training_data[i] for i in file_admet_index]
fit_admet_er_data = [file_er_training_data[i] for i in file_admet_index]

# fit rule er activity
er_rule = math.log(sum([item[0] for item in fit_admet_er_data]) / len(file_admet_index))

rules_fit = []
for step, item in enumerate(fit_admet_er_data):
    if item[0] > er_rule:
        rules_fit.append(step)

fit_rule_data = [fit_admet_data[i] for i in rules_fit]

var_20_data = []
for item in fit_rule_data:
    var_20_data.append([item[i] for i in var_20_names_index])

# write to file
var_20_data = np.array(var_20_data)
f = open('vars_20_value_fit.txt', 'a', encoding='utf-8')
for i in range(20):
    f.write(var_names_20[i]+'\t')
    f.write('start:'+str(min(var_20_data[:, i].tolist()))+'  ')
    f.write('end:'+str(max(var_20_data[:, i].tolist()))+'\n')
print('done')